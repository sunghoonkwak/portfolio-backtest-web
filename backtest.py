import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os

def run_backtest_logic(start_date, end_date, portfolios, selected_tickers=None):
    """
    start_date: 'YYYY-MM-DD'
    end_date: 'YYYY-MM-DD'
    portfolios: dict of { pName: { 'assets': [ {ticker, weight}, ... ], 'interval': 'monthly' } }
    selected_tickers: list of ticker strings to return individual performance for
    """
    if selected_tickers is None: selected_tickers = []
    
    # 1. Gather all tickers, excluding cash-like ones
    all_tickers = set()
    cash_tickers = ['USD', 'KRW', 'CASH']
    for p_content in portfolios.values():
        for asset in p_content['assets']:
            t_upper = asset['ticker'].upper()
            if t_upper not in cash_tickers:
                all_tickers.add(t_upper)
    
    for t in selected_tickers:
        t_upper = t.upper()
        if t_upper not in cash_tickers:
            all_tickers.add(t_upper)
    
    # 2. Download Price Data
    if all_tickers:
        print(f"Downloading data for: {all_tickers}")
        fetch_start = (datetime.strptime(start_date, '%Y-%m-%d') - timedelta(days=10)).strftime('%Y-%m-%d')
        df = yf.download(list(all_tickers), start=fetch_start, end=end_date)
        if df.empty:
            return {"error": f"No price data found for tickers {list(all_tickers)} in the selected period."}
        
        if isinstance(df.columns, pd.MultiIndex):
            has_adj = 'Adj Close' in df.columns.levels[0]
            col_to_use = 'Adj Close' if has_adj else 'Close'
            prices = df[col_to_use]
        else:
            ticker = list(all_tickers)[0]
            col_to_use = 'Adj Close' if 'Adj Close' in df.columns else 'Close'
            if col_to_use in df.columns:
                prices = pd.DataFrame({ticker: df[col_to_use]})
            else:
                return {"error": f"Required price columns not found for {ticker}."}
    else:
        dates = pd.date_range(start=start_date, end=end_date, freq='B')
        prices = pd.DataFrame(index=dates)

    prices = prices.ffill()
    start_dt = pd.to_datetime(start_date)
    prices = prices[prices.index >= start_dt]
    if prices.empty:
        return {"error": "No price data found for the selected period after filtering."}

    results = {
        "portfolios": {},
        "tickers": {}
    }
    total_daily_returns = pd.DataFrame(index=prices.index)
    portfolio_rebal_histories = {}

    freq_map = {'weekly': 'W', 'monthly': 'ME', 'quarterly': 'QE', 'yearly': 'YE'}

    # Portfolio Performance
    for p_name, p_content in portfolios.items():
        assets = p_content['assets']
        interval = p_content.get('interval', 'none').lower()
        
        weights_map = { asset['ticker'].upper(): asset['weight'] for asset in assets }
        
        value = 1.0 # Starting value
        portfolio_values = []
        dates_list = []
        holdings = { t: weights_map[t] * value for t in weights_map }
        
        rebal_dates_set = set()
        if interval != 'none':
            try:
                target_freq = freq_map.get(interval, 'ME')
                rebal_dates_series = prices.groupby(pd.Grouper(freq=target_freq)).apply(lambda x: x.index[-1] if not x.empty else None)
                rebal_dates_set = set(rebal_dates_series.dropna())
            except: pass

        prev_prices = None
        rebal_history = []
        
        for date, current_prices in prices.iterrows():
            if prev_prices is not None:
                new_value = 0
                temp_holdings = {}
                for t in holdings:
                    if t in cash_tickers:
                        curr_h_val = holdings[t]
                    else:
                        ret = current_prices[t] / prev_prices[t] if prev_prices[t] > 0 else 1.0
                        curr_h_val = holdings[t] * ret
                    temp_holdings[t] = curr_h_val
                    new_value += curr_h_val
                
                holdings = temp_holdings
                value = new_value
                
                if date in rebal_dates_set:
                    entry = {"Rebalancing Day": date.strftime('%Y-%m-%d'), "Total Portfolio Value (%)": round(value * 100, 2)}
                    for t in holdings:
                        entry[f"{t} (Before %)"] = round(holdings[t] * 100, 2)
                        holdings[t] = weights_map[t] * value
                        entry[f"{t} (After %)"] = round(holdings[t] * 100, 2)
                    rebal_history.append(entry)

            portfolio_values.append(value)
            dates_list.append(date.strftime('%Y-%m-%d'))
            prev_prices = current_prices

        val_series = pd.Series(portfolio_values, index=prices.index)
        daily_rets = val_series.pct_change().fillna(0)
        total_daily_returns[p_name] = daily_rets
        
        final_return = (portfolio_values[-1] - 1.0) * 100
        std_dev = daily_rets.std() * np.sqrt(252) * 100
        
        results["portfolios"][p_name] = {
            "values": portfolio_values, "dates": dates_list,
            "final_return": final_return, "std_dev": std_dev
        }
        portfolio_rebal_histories[p_name] = rebal_history

    # Ticker Performance
    for t in selected_tickers:
        t_upper = t.upper()
        if t_upper in cash_tickers:
            # Cash return is always 0
            vals = [1.0] * len(prices)
            results["tickers"][t] = {
                "values": vals, "dates": [d.strftime('%Y-%m-%d') for d in prices.index],
                "final_return": 0.0, "std_dev": 0.0
            }
        elif t_upper in prices.columns:
            start_price = prices[t_upper].iloc[0]
            if start_price > 0:
                ticker_values = (prices[t_upper] / start_price).tolist()
                daily_rets = prices[t_upper].pct_change().fillna(0)
                results["tickers"][t] = {
                    "values": ticker_values, "dates": [d.strftime('%Y-%m-%d') for d in prices.index],
                    "final_return": (ticker_values[-1] - 1.0) * 100,
                    "std_dev": daily_rets.std() * np.sqrt(252) * 100
                }

    # Save to Excel
    try:
        output_file = 'backtest_results.xlsx'
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            total_daily_returns.to_excel(writer, sheet_name='Daily Returns')
            for p_name, history in portfolio_rebal_histories.items():
                if history: pd.DataFrame(history).to_excel(writer, sheet_name=p_name, index=False)
                else: pd.DataFrame([{"Info": "No rebalancing"}]).to_excel(writer, sheet_name=p_name, index=False)
        
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        wb = load_workbook(output_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row: cell.font = Font(size=10)
            for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 20
        wb.save(output_file)
    except: pass

    return results
